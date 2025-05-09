import pandas as pd
import math
import os
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook

root = tk.Tk()
root.withdraw()
script_dir = os.path.dirname(os.path.abspath(__file__))
theme_path = os.path.join(script_dir, "themes", "azure.tcl")

try:
    root.tk.call("source", theme_path)
    print("Azure.tcl sourced successfully.")
except tk.TclError as e:
    print(f"Warning: Could not source Azure theme from {theme_path}: {e}")
style = ttk.Style(root)
try:
    style.theme_use("azure-light") 
    print("Azure theme 'azure-light' applied via ttk.Style().theme_use().")
except tk.TclError as e:
    print(f"Failed to apply 'azure-light' theme: {e}. Trying fallback or default.")
    try:
        style.theme_use("clam")
        print("Fell back to 'clam' theme.")
    except tk.TclError:
        print("Could not apply any specific ttk theme. Using system default.")

def load_employees_from_excel(file_path):
    """Load employees from an Excel file."""
    if not os.path.exists(file_path):
        return []

    try:
        df = pd.read_excel(file_path, keep_default_na=False, na_values=['NaN', 'nan', '', '#N/A', 'N/A'])
    except Exception as e:
        print(f"Error loading Excel file {file_path}: {e}")
        messagebox.showerror("Excel Load Error", f"Could not load {os.path.basename(file_path)}.\nError: {e}")
        return []


    employees = []
    for index, row in df.iterrows(): # Added index for better error reporting
        try:
            name = row.get('Name', '') # Use .get with a default
            if pd.isna(name) or not isinstance(name, str) or name.strip() == "":
                # print(f"Warning: Skipping row {index+2} in {os.path.basename(file_path)} due to missing or invalid Name: '{name}'")
                continue # Skip rows with no valid name
            name = str(name).strip()


            availability = {
                'Sun': str(row.get('Sun', '')).strip().lower() == "yes",
                'Mon': str(row.get('Mon', '')).strip().lower() == "yes",
                'Tue': str(row.get('Tue', '')).strip().lower() == "yes",
                'Wed': str(row.get('Wed', '')).strip().lower() == "yes",
                'Thu': str(row.get('Thu', '')).strip().lower() == "yes",
                'Fri': str(row.get('Fri', '')).strip().lower() == "yes",
                'Sat': str(row.get('Sat', '')).strip().lower() == "yes"
            }

            notes_val = row.get('Notes', '') # Default to empty string if 'Notes' column is missing
            
            # Explicitly handle cases where notes might be NaN (float) or other non-string types
            if pd.isna(notes_val): # Checks for pandas NaN, numpy.nan, None
                notes = ""
            elif isinstance(notes_val, float): # e.g. if a number was in notes somehow
                notes = str(int(notes_val)) if notes_val.is_integer() else str(notes_val)
            elif notes_val is None: # Handles explicit None
                notes = ""
            else:
                notes = str(notes_val).strip() # Convert to string and strip whitespace

            # The debug print can be noisy, consider commenting out for normal use
            # print(f"Loaded Employee: {name}, Availability: {availability}, Notes: '{notes}'") 
            employees.append(Employee(name, availability, notes))
        except Exception as e:
            print(f"Error processing row {index+2} in {os.path.basename(file_path)}: {row.to_dict()}. Error: {e}")
            # Decide if you want to skip the row or raise the error

    return employees
class Employee:
    def __init__(self, name, availability, notes=""):
        self.name = name
        self.availability = availability  # availability as a dictionary {'Sun': True, 'Mon': False, ...}
        self.notes = notes

    def __str__(self):
        return f"Employee(name={self.name}, availability={self.availability}, notes={self.notes})"

class Schedule:
    def __init__(self, days, employees):
        self.days = days
        self.employees = employees  # Store employees in the class
        self.employees_needed = {day: [] for day in days}  # Initialize employees_needed
        self.schedule = {day: [] for day in days}
        self.unassigned_employees = {day: [] for day in days}  # Track unassigned employees per day
        self.final_schedule = {day: [] for day in days}

    def set_employees_needed(self, employees_needed):
        """Set the number of employees needed for each day."""
        self.employees_needed = employees_needed  # Update the employees_needed attribute

    def get_max_employees_for_day(self, day):
        """Get the maximum number of employees needed for a specific day."""
        return self.employees_needed.get(day, 0)

    def add_employee_to_day(self, day, employee, force=False):
        """Assign an employee to a day, considering availability and employee limits."""
        if day not in self.schedule:
            print(f"Invalid day: {day}")
            return

        # Check availability unless force is True
        if not employee.availability.get(day, False) and not force:
            print(f"{employee.name} is not available on {day}.")
            return

        # Only check the max employee limit if not forcing a manual assignment
        if True:
            self.schedule[day].append(employee)
            self.schedule[day] = sorted(self.schedule[day], key=lambda emp: emp.name)
            print(f"Assigned {employee.name} to {day} and sorted alphabetically.")
            
            # Remove the employee from the unassigned list if they were unassigned
            if employee in self.unassigned_employees[day]:
                self.unassigned_employees[day].remove(employee)
                print(f"Removed {employee.name} from unassigned employees for {day}.")

    def generate_schedule(self):
        self.unassigned_employees = {day: [] for day in self.days}  # Reset unassigned employees
        self.schedule = {day: [] for day in self.days}  # Clear the existing schedule

        for day in self.days:
            needed = self.get_max_employees_for_day(day)  # Get the number of needed employees for the day

            # Filter available employees for the current day
            available_employees = [emp for emp in self.employees if emp.availability.get(day, False)]

            # Ensure employee names are valid strings for sorting, handle potential errors
            available_employees = [emp for emp in available_employees if isinstance(emp.name, str)]

            # Reverse order for Saturday and Sunday
            if day in ['Sun', 'Sat']:
                available_employees.reverse()  # Reverse the list of employees for these days

            print(f"Available employees for {day}: {[emp.name for emp in available_employees]}")  # Debug output

            assigned_count = len(self.schedule[day])  # Start with already assigned count
            
            for employee in available_employees:
                print(f"Assigning {employee.name} to {day}.")  # Debug output
                self.add_employee_to_day(day, employee)

            # Collect unassigned employees
            for employee in available_employees:
                if employee not in self.schedule[day]:
                    self.unassigned_employees[day].append(employee)

            print(f"Assigned employees for {day}: {[emp.name for emp in self.schedule[day]]}")  # Debug output

        self.refresh_unassigned_employees()
        return self.unassigned_employees

    def print_schedule(self):
        output = "Final Schedule:\n"
        for day, employees in self.schedule.items():
            employee_names = sorted([emp.name for emp in employees])
            output += f"{day}: {', '.join(employee_names) if employee_names else 'No employees assigned'}\n"
        
        # Add unassigned employees
        output += "\nUnassigned Employees:\n"
        for day, unassigned in self.unassigned_employees.items():
            # Extract names of unassigned employees and sort alphabetically
            unassigned_names = sorted([emp.name for emp in unassigned])
            output += f"{day}: {', '.join(unassigned_names) if unassigned_names else 'All employees assigned'}\n"

        return output

    def manually_add_employee(self, day, employee, force=True):
        """Manually add an employee to a day, considering availability."""
        if day not in self.schedule:
            return
        if not employee.availability.get(day, False) and not force:
            return
        self.add_employee_to_day(day, employee, force)

    def refresh_unassigned_employees(self):
        """Refresh the unassigned employees list and sort them alphabetically."""
        self.unassigned_employees = {
            day: sorted(
                [emp for emp in self.employees if emp not in self.schedule[day]], 
                key=lambda emp: str(emp.name)  # Ensure name is treated as a string
            )
            for day in self.days
        }
        print("Unassigned employees refreshed.")
class ScheduleWindow:
    def __init__(self, master):
        # Master window
        self.master = master
        self.master.deiconify()
        self.master.title("Schedule Management")
        self.master.rowconfigure(0, weight=1)
        self.master.columnconfigure(0, weight=1)

        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.app_container = ttk.Frame(self.master, padding="0") # Use ttk.Frame
        self.app_container.grid(row=0, column=0, sticky="nsew")
        self.app_container.columnconfigure(0, weight=1)
        self.app_container.rowconfigure(0, weight=0)
        self.app_container.rowconfigure(1, weight=1)

        self.drag_data = None
        self.finalized_employees = []
        self.all_possible_employees = []
        self.days = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        self.employees = []
        self.schedule = Schedule(days=self.days, employees=[])

        # Top frame for buttons
        top_frame = ttk.Frame(self.app_container)
        top_frame.grid(row=0, column=0, columnspan=5, sticky="ew", padx=5, pady=5)
        top_frame.columnconfigure(0, weight=0)
        top_frame.columnconfigure(1, weight=1)

        # Dropdown Set Schedule
        self.set_schedule_dropdown_label = ttk.Label(top_frame, text="Set Schedule List:", font=("Arial", 10, "bold"))
        self.set_schedule_dropdown_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.file_selection_set = ttk.Combobox(top_frame, values=self.get_excel_files(), state='readonly', width=48)
        self.file_selection_set.set("Select Set Employee Excel Sheet")
        self.file_selection_set.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.file_selection_set.bind("<<ComboboxSelected>>", self.on_file_selected_set)

        # Dropdown Unset Schedule
        self.unset_schedule_dropdown_label = ttk.Label(top_frame, text="Unset Schedule List:", font=("Arial", 10, "bold"))
        self.unset_schedule_dropdown_label.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.file_selection_unset = ttk.Combobox(top_frame, values=self.get_excel_files(), state='readonly', width=48)
        self.file_selection_unset.set("Select Unset Employee Excel Sheet")
        self.file_selection_unset.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        self.file_selection_unset.bind("<<ComboboxSelected>>", self.on_file_selected_unset)

        # Buttons
        self.push_to_final_button = ttk.Button(top_frame, text="Push Changes To Final Schedule", command=self.push_changes_to_final_schedule)
        self.push_to_final_button.grid(row=0, column=3, rowspan=2, padx=5, pady=5)

        self.employee_manager_button = ttk.Button(top_frame, text="Manage Employees", command=self.open_manage_employees_window)
        self.employee_manager_button.grid(row=0, column=4, rowspan=2, padx=5, pady=5)

        toggle_btn = ttk.Button(top_frame, text="Toggle Theme", command=self.toggle_theme)
        toggle_btn.grid(row=0, column=5, rowspan=2, padx=5, pady=5)

        # Main frame
        main_frame = ttk.Frame(self.app_container)
        main_frame.grid(row=1, column=0, sticky="nsew")
        for i in range(2): main_frame.columnconfigure(i, weight=1)
        main_frame.rowconfigure(0, weight=1)  # Schedule and Finalized Treeviews
        main_frame.rowconfigure(1, weight=1)  # Unassigned and Employee Treeviews
        main_frame.rowconfigure(2, weight=0)  # Set Employees Needed stays small

        # Unassigned Treeview
        unassigned_frame = ttk.Frame(main_frame)
        unassigned_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)

        unassigned_scroll = ttk.Scrollbar(unassigned_frame)
        unassigned_scroll.grid(row=0, column=1, sticky="ns")

        self.unassigned_tree = ttk.Treeview(unassigned_frame, columns=["Row"] + self.days, show="headings", height=15, yscrollcommand=unassigned_scroll.set)
        self.unassigned_tree.grid(row=0, column=0, sticky="nsew")
        unassigned_scroll.config(command=self.unassigned_tree.yview)

        unassigned_frame.columnconfigure(0, weight=1)
        unassigned_frame.rowconfigure(0, weight=1)

        self.unassigned_tree.heading("Row", text="#")
        self.unassigned_tree.column("Row", width=30, anchor="center", stretch=True)
        for day in self.days:
            self.unassigned_tree.heading(day, text=day)
            self.unassigned_tree.column(day, width=80, anchor="center", stretch=True)
        self.unassigned_tree.tag_configure('oddrow', background='#AAC1DC')
        self.unassigned_tree.tag_configure('evenrow', background='#ffffff')
        self.unassigned_tree.bind("<Double-1>", self.on_unassigned_double_click)

        # Drag and drop bindings
        self.unassigned_tree.bind("<ButtonPress-1>", self.on_drag_start_unassigned)
        self.master.bind("<ButtonRelease-1>", self.on_drag_release_anywhere, add=True) # Global release

        # Finalized Treeview
        finalized_frame = ttk.Frame(main_frame)
        finalized_frame.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=5, pady=5)

        finalized_scroll = ttk.Scrollbar(finalized_frame)
        finalized_scroll.grid(row=0, column=1, sticky="ns")

        self.finalized_tree = ttk.Treeview(finalized_frame, columns=["Row"] + self.days, show="headings", height=15, yscrollcommand=finalized_scroll.set)
        self.finalized_tree.grid(row=0, column=0, sticky="nsew")
        finalized_scroll.config(command=self.finalized_tree.yview)

        finalized_frame.columnconfigure(0, weight=1)
        finalized_frame.rowconfigure(0, weight=1)

        self.finalized_tree.heading("Row", text="#")
        self.finalized_tree.column("Row", width=30, anchor="center", stretch=True)
        for day in self.days:
            self.finalized_tree.heading(day, text=day)
            self.finalized_tree.column(day, width=80, anchor="center", stretch=True)
        self.finalized_tree.tag_configure('oddrow', background='#AAC1DC')
        self.finalized_tree.tag_configure('evenrow', background='#ffffff')
        self.finalized_tree.tag_configure('highlight', background='lightblue')
        self.finalized_tree.bind("<Double-1>", self.on_final_schedule_double_click)
        self.finalized_tree.bind("<Button-1>", self.on_final_schedule_single_click)

        # Employee List Treeview
        employee_list_frame = ttk.Frame(main_frame)
        employee_list_frame.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)

        employee_scroll = ttk.Scrollbar(employee_list_frame)
        employee_scroll.grid(row=0, column=1, sticky="ns")

        self.employee_list_tree = ttk.Treeview(employee_list_frame, columns=["Name", "Availability", "Notes"], show="headings", height=15, yscrollcommand=employee_scroll.set)
        self.employee_list_tree.grid(row=0, column=0, sticky="nsew")
        employee_scroll.config(command=self.employee_list_tree.yview)

        employee_list_frame.columnconfigure(0, weight=1)
        employee_list_frame.rowconfigure(0, weight=1)

        self.employee_list_tree.heading("Name", text="Name")
        self.employee_list_tree.column("Name", width=200, anchor="w", stretch=True)
        self.employee_list_tree.heading("Availability", text="Available Days")
        self.employee_list_tree.column("Availability", width=200, anchor="w", stretch=True)
        self.employee_list_tree.heading("Notes", text="Notes")
        self.employee_list_tree.column("Notes", width=200, anchor="w", stretch=True)
        self.employee_list_tree.tag_configure('oddrow', background='#AAC1DC')
        self.employee_list_tree.tag_configure('evenrow', background='#ffffff')

        self.employee_list_tree.bind("<ButtonPress-1>", self.on_drag_start_employee_list)

        # Set Employees Needed Frame
        set_employees_needed_frame = ttk.Frame(main_frame)
        set_employees_needed_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=5)

        # Set Employees Needed Label
        set_employees_needed_label = ttk.Label(set_employees_needed_frame, text="Set Employees Needed", font=("Arial", 12, "bold"))
        set_employees_needed_label.grid(row=0, column=0, columnspan=15, rowspan=2, sticky="w", pady=(0, 35))

        # Entries for Each Day (placed in two rows if needed)
        self.employees_needed_entries = {}
        for idx, day in enumerate(self.days):
            col = idx % 7
            row = 1 if idx < 7 else 2
            label = ttk.Label(set_employees_needed_frame, text=f"{day}:")
            label.grid(row=row, column=col*2, sticky="w", padx=(0, 2), pady=5)
            entry = ttk.Entry(set_employees_needed_frame, width=5)
            entry.grid(row=row, column=col*2 + 1, sticky="w", padx=(0, 5), pady=5)
            self.employees_needed_entries[day] = entry

        # Submit Button aligned under the entry fields
        self.submit_needed_button = ttk.Button(set_employees_needed_frame, text="Submit", command=self.submit_employees_needed)
        self.submit_needed_button.grid(row=1, column=14, rowspan=2, sticky="w", padx=10)

        # Employee info frame
        employee_info_frame = ttk.Frame(main_frame)
        employee_info_frame.grid(row=2, column=1, sticky="w", padx=5, pady=5)

        # Name entry (left top)
        name_label = ttk.Label(employee_info_frame, text="Name:", font=("Arial", 12))
        name_label.grid(row=0, column=0, sticky="sw", padx=5, pady=(40, 0))
        name_entry = ttk.Entry(employee_info_frame, width=30)
        name_entry.grid(row=0, column=1, sticky="sw", padx=5, pady=(40, 0))

        # Checkboxes (same row below name, iterate to the right)
        check_vars = {day: tk.BooleanVar() for day in self.days}
        check_frame = ttk.Frame(employee_info_frame)
        check_frame.grid(row=1, column=0, columnspan=2, sticky="sw", padx=5, pady=(0, 40))
        for i, day in enumerate(self.days):
            tk.Checkbutton(check_frame, text=day, variable=check_vars[day]).grid(row=0, column=i, sticky="w", padx=2)

        # Notes entry (right)
        notes_label = ttk.Label(employee_info_frame, text="Notes:", font=("Arial", 12))
        notes_label.grid(row=0, column=2, sticky="ew", padx=(20, 40), pady=0)
        notes_entry = tk.Text(employee_info_frame, width=50, height=4, wrap="word")
        notes_entry.grid(row=1, column=2, sticky="ew", padx=(20, 40), pady=(0,25))
        notes_entry.insert("1.0", "Edit note here")

        self.populate_employee_list()
        if not self.employees: # If on_file_selected_unset hasn't run yet
            self.employees = list(self.all_possible_employees) # Make a copy
            self.schedule.employees = self.employees
        self.refresh_employee_list_tree()
        self.refresh_unassigned_employees() 
        self.refresh_finalized_schedule()

    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.master.destroy()

    def toggle_theme(self):
        # This now uses the Tcl proc 'set_theme' defined in azure.tcl
        try:
            s = ttk.Style()
            current_ttk_theme = s.theme_use()
            
            print(f"Current Ttk theme for toggle: {current_ttk_theme}")

            if "dark" in current_ttk_theme.lower():
                self.master.tk.call("set_theme", "light")
                print("Called Tcl proc: set_theme light")
            else:
                self.master.tk.call("set_theme", "dark")
                print("Called Tcl proc: set_theme dark")
        except tk.TclError as e:
            messagebox.showerror("Theme Error", f"Could not switch theme via Tcl 'set_theme'.\nError: {e}")
            print(f"Error toggling theme via Tcl: {e}")

    def on_drag_start_unassigned(self, event):
        item_id = self.unassigned_tree.identify_row(event.y)
        column_id_str = self.unassigned_tree.identify_column(event.x)

        if not item_id or not column_id_str:
            return

        column_index_tree = int(column_id_str.replace('#', ''))
        if column_index_tree <= 1:
            return

        item_values = self.unassigned_tree.item(item_id, 'values')
        employee_name_index_in_item_values = column_index_tree - 1

        if not item_values or employee_name_index_in_item_values >= len(item_values):
            return

        employee_name = item_values[employee_name_index_in_item_values]
        if not employee_name:
            return

        selected_employee = None
        # The unassigned_tree displays names from self.employees filtered by final_schedule
        for emp_obj in self.employees: # Search in the master list of ScheduleWindow
            if emp_obj.name == employee_name:
                selected_employee = emp_obj
                break
        
        if selected_employee:
            self.drag_data = {'employee': selected_employee, 'source_widget_type': 'unassigned_tree'}
            self.master.config(cursor="hand2")
        else:
            print(f"Error: Employee object for '{employee_name}' (dragged from unassigned_tree) "
                    f"not found in self.employees. Sync issue likely.")

    def on_drag_start_employee_list(self, event):
        item_id = self.employee_list_tree.identify_row(event.y)
        if not item_id:
            return

        item_values = self.employee_list_tree.item(item_id, 'values')
        if not item_values or not item_values[0]: 
            return

        employee_name = item_values[0]
        selected_employee_object = None
        # Source from the static list
        for emp in self.all_possible_employees: 
            if emp.name == employee_name:
                selected_employee_object = emp
                break
        
        if selected_employee_object:
            self.drag_data = {
                'employee': selected_employee_object, # This is an Employee object from all_possible_employees
                'source_widget_type': 'employee_list_tree'
            }
            self.master.config(cursor="hand2")
        else:
            print(f"Error: Could not find employee object for '{employee_name}' "
                    f"from employee_list_tree in self.all_possible_employees.")

    def on_drag_release_anywhere(self, event):
        if not self.drag_data: 
            return

        employee_to_assign = self.drag_data.get('employee')
        source_type = self.drag_data.get('source_widget_type') # 'unassigned_tree' or 'employee_list_tree'
        
        current_cursor = self.master.cget("cursor")
        self.drag_data = None 
        if current_cursor == "hand2": 
            self.master.config(cursor="")

        if not employee_to_assign or source_type not in ['unassigned_tree', 'employee_list_tree']:
            # print(f"Drag release: No valid employee or source type. Source: {source_type}")
            return

        x_root, y_root = event.x_root, event.y_root
        target_widget = self.master.winfo_containing(x_root, y_root)

        if target_widget == self.finalized_tree:
            try:
                tree_x = target_widget.winfo_pointerx() - target_widget.winfo_rootx()
            except tk.TclError: 
                return

            target_column_id_str = self.finalized_tree.identify_column(tree_x)
            if not target_column_id_str: return

            target_column_index_tree = int(target_column_id_str.replace('#', '')) 
            if target_column_index_tree <= 1: 
                return
            
            day_list_index = target_column_index_tree - 2 
            if 0 <= day_list_index < len(self.days):
                target_day = self.days[day_list_index]
                
                if not employee_to_assign.availability.get(target_day, False):
                    if not messagebox.askyesno("Not Available", 
                                                f"{employee_to_assign.name} is not normally available on {target_day}. "
                                                "Assign anyway?"):
                        return 

                self.assign_employee_to_final_schedule(employee_to_assign, target_day, method=f"drag from {source_type}")
        # else:
            # print(f"Drag released, but not on finalized_tree. Target: {target_widget}, Source: {source_type}")

    def assign_employee_to_final_schedule(self, employee_dragged, day, method="unknown"):
        if not isinstance(employee_dragged, Employee):
            messagebox.showerror("Error", "Invalid employee data for assignment.")
            return

        # 1. Get the canonical Employee object from self.all_possible_employees.
        # This ensures we're always working with the definitive version of the employee from the static list.
        canonical_employee_object = None
        for emp_in_static_list in self.all_possible_employees:
            if emp_in_static_list.name == employee_dragged.name:
                canonical_employee_object = emp_in_static_list
                break
        
        if not canonical_employee_object:
            messagebox.showerror("Error", f"Critical: Employee {employee_dragged.name} not found in the master directory (all_possible_employees).")
            return
        
        # Use this canonical object for all further operations in this function.
        employee_to_assign = canonical_employee_object

        # 2. Check if this employee is in the current ACTIVE pool (self.employees).
        # If not, add them. This is crucial for consistency, especially if they were
        # dragged from employee_list_tree and weren't part of the loaded "Unset" list.
        found_in_active_pool = False
        for active_emp in self.employees:
            if active_emp.name == employee_to_assign.name:
                # If found, ensure we use the instance from the active pool for schedule.add_employee_to_day
                # if that method relies on object identity from self.schedule.employees.
                # However, since we're using canonical_employee_object, and we just added it if missing,
                # this re-assignment might be redundant if add_employee_to_day only cares about name/availability.
                # For safety, let's assume we should use the object that's now definitively in self.employees.
                employee_to_assign_for_schedule_class = active_emp 
                found_in_active_pool = True
                break

        if not found_in_active_pool:
            self.employees.append(employee_to_assign) # Add the canonical object
            self.schedule.employees = self.employees # Keep Schedule class's list in sync
            employee_to_assign_for_schedule_class = employee_to_assign # It's now in self.employees
            print(f"Added {employee_to_assign.name} to active pool (self.employees) as they were assigned from static list.")
        
        # 3. Add to staging schedule (self.schedule.schedule)
        # Pass the object that is confirmed to be in (or just added to) self.employees
        self.schedule.add_employee_to_day(day, employee_to_assign_for_schedule_class, force=True) 

        # 4. Add to final schedule (self.schedule.final_schedule)
        if day not in self.schedule.final_schedule: self.schedule.final_schedule[day] = []
        
        # Check against the final schedule using the canonical object.
        is_already_in_final = any(final_emp.name == employee_to_assign.name for final_emp in self.schedule.final_schedule[day])

        if not is_already_in_final:
            self.schedule.final_schedule[day].append(employee_to_assign) # Add the canonical object
            self.schedule.final_schedule[day] = sorted(
                self.schedule.final_schedule[day],
                key=lambda emp_obj: emp_obj.name.lower()
            )
        else:
            messagebox.showinfo("Already Assigned", f"{employee_to_assign.name} is already in the final schedule for {day}.")
        
        # 5. Refresh UIs
        self.refresh_unassigned_employees() 
        self.refresh_finalized_schedule()

    def on_unassigned_double_click(self, event):
        item_id = self.unassigned_tree.identify_row(event.y)
        column_id_str = self.unassigned_tree.identify_column(event.x)

        if not item_id or not column_id_str: return

        column_index_tree = int(column_id_str.replace('#', ''))
        if column_index_tree <= 1: return 

        day_list_index = column_index_tree - 2 
        if not (0 <= day_list_index < len(self.days)): return
        selected_day = self.days[day_list_index]
        
        item_values = self.unassigned_tree.item(item_id, 'values')
        employee_name_index_in_item_values = column_index_tree -1

        if not item_values or employee_name_index_in_item_values >= len(item_values): return

        selected_employee_name = item_values[employee_name_index_in_item_values]
        if not selected_employee_name: return
        
        selected_employee_object = next((emp for emp in self.employees if emp.name == selected_employee_name), None)
        
        if not selected_employee_object:
            messagebox.showerror("Error", f"Employee '{selected_employee_name}' (from unassigned_tree) not found in master list.")
            return

        if not selected_employee_object.availability.get(selected_day, False):
            if not messagebox.askyesno("Not Available", 
                                        f"{selected_employee_object.name} is not normally available on {selected_day}. "
                                        "Assign anyway?"):
                return

        if messagebox.askyesno("Manual Assignment", f"Do you want to manually add {selected_employee_object.name} to {selected_day}?"):
            self.assign_employee_to_final_schedule(selected_employee_object, selected_day, method="double-click from unassigned")

    def on_final_schedule_double_click(self, event):
        """Handle double-clicking on any employee in the final schedule to remove them."""
        item_id = self.finalized_tree.identify_row(event.y)
        if not item_id:
            return  # No row selected

        column = self.finalized_tree.identify_column(event.x)
        column_index = int(column.replace('#', '')) - 1

        item_values = self.finalized_tree.item(item_id, 'values')

        if column_index == 0:
            return  # Ignore row number column

        selected_employee = item_values[column_index]
        if not selected_employee:
            return  # Empty cell

        selected_day = self.days[column_index - 1]

        confirm = messagebox.askyesno("Remove Employee", f"Do you want to remove {selected_employee} from {selected_day} in the final schedule?")

        if confirm:
            # Find and remove the employee object
            final_employees = self.schedule.final_schedule[selected_day]
            self.schedule.final_schedule[selected_day] = [
                emp for emp in final_employees if emp.name != selected_employee
            ]

            self.refresh_finalized_schedule()

    def on_final_schedule_single_click(self, event):
        # Determine base tag for alternating colors before applying highlight
        children = self.finalized_tree.get_children()
        base_tags = {child_id: ('oddrow' if i % 2 == 0 else 'evenrow') for i, child_id in enumerate(children)}

        for item_id in children:
            self.finalized_tree.item(item_id, tags=(base_tags[item_id],)) # Reset to base tag

        item_id = self.finalized_tree.identify_row(event.y)
        column_str = self.finalized_tree.identify_column(event.x)

        if not item_id or column_str == '#0': return # Click on header or empty space

        column_index_tree = int(column_str.replace('#', '')) 
        if column_index_tree <= 1: return # Click on row number column

        item_values = self.finalized_tree.item(item_id, 'values')
        actual_data_column_index = column_index_tree - 1

        if actual_data_column_index >= len(item_values): return # Click beyond data columns
        
        selected_employee_name = item_values[actual_data_column_index]
        if not selected_employee_name: return

        for child_id in children:
            current_values = self.finalized_tree.item(child_id, 'values')
            # Check if employee name is in any day column for this row (values[1] onwards)
            if any(name == selected_employee_name for name in current_values[1:]):
                # Append 'highlight' to existing base tag
                new_tags = (base_tags[child_id], 'highlight')
                self.finalized_tree.item(child_id, tags=new_tags)

    def get_excel_files(self):
        """Retrieve all available Excel files for selection."""
        excel_dir = os.path.join(os.getcwd(), "excel_sheets")  # Assuming excel files are in an 'excel_sheets' folder
        return [f for f in os.listdir(excel_dir) if f.endswith('.xlsx')]

    def on_file_selected_set(self, event):
        """Load employees from the selected Excel file and populate final schedule."""
        selected_file = self.file_selection_set.get()
        if selected_file and selected_file != "Select Set Employee Excel Sheet":
            excel_dir = os.path.join(os.getcwd(), "excel_sheets")
            file_path = os.path.join(excel_dir, selected_file)

            self.set_employees = load_employees_from_excel(file_path)
            if self.set_employees:
                print(f"[SET] Loaded employees from {file_path}: {[emp.name for emp in self.set_employees]}")
                self.schedule.final_schedule = {day: [] for day in self.days}

                for emp in self.set_employees:
                    for day in self.days:
                        if emp.availability.get(day, False):
                            self.schedule.final_schedule[day].append(emp)

                self.refresh_finalized_schedule()
            else:
                print(f"[SET] No employees loaded from {file_path}.")

    def on_file_selected_unset(self, event):
        selected_file = self.file_selection_unset.get()
        
        if selected_file and selected_file != "Select Unset Employee Excel Sheet":
            excel_dir = os.path.join(os.getcwd(), "excel_sheets")
            file_path = os.path.join(excel_dir, selected_file)

            self.employees = load_employees_from_excel(file_path) # Update ACTIVE pool
            self.schedule.employees = self.employees # Update Schedule object's active pool
            
            if self.employees:
                print(f"[UNSET] Loaded {len(self.employees)} employees into ACTIVE pool (self.employees) from {file_path}")
            else:
                print(f"[UNSET] No employees loaded from {file_path}. ACTIVE pool (self.employees) is now empty.")
        else:
            # Fallback if "Select Unset..." is chosen or dropdown is cleared
            # Option 1: Make active pool empty
            # self.employees = []
            # Option 2: Default to all_possible_employees
            self.employees = list(self.all_possible_employees) # Make a copy
            self.schedule.employees = self.employees
            print("[UNSET] No specific file selected. ACTIVE pool (self.employees) defaulted to all possible employees.")

        self.schedule.schedule = {day: [] for day in self.days} # Clear staging schedule

        # self.populate_employee_list() # NO LONGER NEEDED HERE for the static tree
        self.refresh_unassigned_employees() # Refreshes unassigned_tree based on new ACTIVE self.employees vs final_schedule
        # self.refresh_finalized_schedule() # Generally not directly affected but good for consistency

    def push_changes_to_final_schedule(self):
        """Push assigned employees to final schedule without overwriting existing entries."""
        for day in self.days:
            unassigned_employees = self.schedule.unassigned_employees[day]
            final_employees = self.schedule.final_schedule[day]

            # Only add new employees who are not already in the final list
            for emp in unassigned_employees:
                if emp not in final_employees:
                    final_employees.append(emp)

            # Sort in-place after adding
            self.schedule.final_schedule[day] = sorted(final_employees, key=lambda emp: emp.name.lower())

        # Refresh the finalized schedule treeview
        self.refresh_finalized_schedule()

    def refresh_unassigned_employees(self):
        self.unassigned_tree.delete(*self.unassigned_tree.get_children())

        day_to_unassigned_objects = {} 
        for day_key in self.days:
            assigned_in_final_for_day_names = {emp.name for emp in self.schedule.final_schedule.get(day_key, [])}
            
            unassigned_for_this_day = []
            # Iterate through the ACTIVE pool (self.employees)
            for emp_candidate in self.employees: 
                if emp_candidate.availability.get(day_key, False) and \
                    emp_candidate.name not in assigned_in_final_for_day_names:
                    unassigned_for_this_day.append(emp_candidate)
            
            day_to_unassigned_objects[day_key] = sorted(unassigned_for_this_day, key=lambda e: e.name.lower())

        max_rows = max((len(emp_list) for emp_list in day_to_unassigned_objects.values()), default=0)
        for i in range(max_rows):
            row_values = [i + 1] 
            for day_k in self.days:
                if i < len(day_to_unassigned_objects[day_k]):
                    row_values.append(day_to_unassigned_objects[day_k][i].name)
                else:
                    row_values.append('')
            tag = 'oddrow' if i % 2 == 0 else 'evenrow'
            self.unassigned_tree.insert('', 'end', values=row_values, tags=(tag,))
        # print("Unassigned employees treeview refreshed.")

    def refresh_finalized_schedule(self):
        """Display current finalized assignments per day (no max_employees logic)."""
        self.finalized_tree.delete(*self.finalized_tree.get_children())
        max_rows = max((len(self.schedule.final_schedule[day]) for day in self.days), default=0)

        for i in range(max_rows):
            row = []
            for day in self.days:
                employees = sorted(self.schedule.final_schedule[day], key=lambda emp: emp.name.lower())
                if i < len(employees):
                    row.append(employees[i].name)
                else:
                    row.append('')
            tag = 'oddrow' if i % 2 == 0 else 'evenrow'
            self.finalized_tree.insert('', 'end', values=[i + 1] + row, tags=(tag,))

    def populate_employee_list(self):
        """Loads employees from "Employees_Full_List.xlsx" into self.all_possible_employees.
        This list is static and primarily for the employee_list_tree."""
        file_path = os.path.join(os.getcwd(), "excel_sheets", "Employees_Full_List.xlsx")
        if not os.path.exists(file_path):
            print("Default 'Employees_Full_List.xlsx' not found.")
            self.all_possible_employees = []
            return

        self.all_possible_employees = load_employees_from_excel(file_path)
        print(f"Loaded {len(self.all_possible_employees)} employees into self.all_possible_employees (static list).")

    def refresh_employee_list_tree(self):
        """
        Populates the employee_list_tree (bottom right) from self.all_possible_employees.
        """
        if not hasattr(self, 'employee_list_tree'): return
        for row in self.employee_list_tree.get_children():
            self.employee_list_tree.delete(row)
        
        # Display whatever is currently in self.all_possible_employees, sorted
        sorted_static_employees = sorted(self.all_possible_employees, key=lambda emp: emp.name.lower())

        for index, emp in enumerate(sorted_static_employees):
            available_days_str = ', '.join(day for day in self.days if emp.availability.get(day, False))
            notes_str = emp.notes or ""
            tag = 'oddrow' if index % 2 == 0 else 'evenrow'
            self.employee_list_tree.insert("", "end", values=(emp.name, available_days_str, notes_str), tags=(tag,))
        # print(f"Employee list tree (static) refreshed with {len(sorted_static_employees)} employees.")

    def submit_employees_needed(self):
        """Submit the employees needed for each day."""
        success_messages = []
        error_messages = []

        for day, entry in self.employees_needed_entries.items():
            try:
                employees_needed = int(entry.get())
                self.schedule.employees_needed[day] = employees_needed
                success_messages.append(f"Set employees needed for {day} to {employees_needed}.")
            except ValueError:
                error_messages.append(f"Invalid number for {day}. Please enter a valid integer.")

        # Display a single message for all success messages
        if success_messages:
            success_message = "\n".join(success_messages)
            messagebox.showinfo("Success", success_message)

        # Display a single message for all error messages
        if error_messages:
            error_message = "\n".join(error_messages)
            messagebox.showerror("Error", error_message)

    def open_manage_employees_window(self):
        excel_dir = os.path.join(os.getcwd(), "excel_sheets")
        file_to_manage_name = "Employees_Full_List.xlsx"
        print(f"Manage Employees window will operate on: {file_to_manage_name}") # For debugging

        file_path = os.path.join(excel_dir, file_to_manage_name)
        if not os.path.exists(file_path):
            if messagebox.askyesno("File Not Found", f"The master employee file '{file_to_manage_name}' was not found.\n"
                                    "Would you like to create it with default headers?"):
                wb_new = Workbook()
                header = ["Name"] + self.days + ["Notes"]
                try:
                    wb_new.active.append(header)
                    wb_new.save(file_path)
                    print(f"Created '{file_path}'")
                except Exception as e:
                    messagebox.showerror("File Creation Error", f"Could not create '{file_path}'.\nError: {e}")
                    return
            else:
                return # User chose not to create the file

        deleted_file_path = os.path.join(excel_dir, "recently_deleted.xlsx")
        if not os.path.exists(deleted_file_path):
            wb_del = Workbook()
            header_del = ["Name"] + self.days + ["Notes"] # Match header for consistency
            try:
                wb_del.active.append(header_del)
                wb_del.save(deleted_file_path)
            except Exception as e:
                print(f"Could not create recently_deleted.xlsx: {e}")
                # Non-critical, so we can continue

        top = tk.Toplevel(self.master)
        top.title("Manage Employees")
        top.geometry("1300x700")

        # allow resizing
        top.rowconfigure(1, weight=1)
        top.columnconfigure(0, weight=1)

        wb = load_workbook(file_path)
        sheet = wb.active

        employee_row_map = {}
        left_tree = None
        right_tree = None
        left_employees = []
        right_employees = []

        # Treeview frame (row 0)
        tree_frame = tk.Frame(top)
        tree_frame.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)
        top.rowconfigure(0, weight=1)
        top.columnconfigure(0, weight=1)
        top.columnconfigure(1, weight=1)

        # Left frame for left tree and its scrollbar
        left_frame = tk.Frame(tree_frame)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)  # Increased weight for more expansion room
        left_frame.rowconfigure(0, weight=1)
        left_frame.columnconfigure(0, weight=1)

        left_tree = ttk.Treeview(left_frame, columns=("Name", "Availability", "Notes"), show="headings")
        left_tree.heading("Name", text="Name")
        left_tree.column("Name", width=200, anchor="w", stretch=True)
        left_tree.heading("Notes", text="Notes")
        left_tree.column("Notes", width=200, anchor="w", stretch=True)
        left_tree.heading("Availability", text="Available Days")
        left_tree.column("Availability", width=200, anchor="w", stretch=True)
        left_tree.grid(row=0, column=0, sticky="nsew")

        left_scroll = ttk.Scrollbar(left_frame, orient="vertical", command=left_tree.yview)
        left_tree.configure(yscrollcommand=left_scroll.set)
        left_scroll.grid(row=0, column=1, sticky="ns")

        # Right frame for right tree and its scrollbar
        right_frame = tk.Frame(tree_frame)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        tree_frame.columnconfigure(1, weight=1)  # Right frame weight remains 1
        right_frame.rowconfigure(0, weight=1)
        right_frame.columnconfigure(0, weight=1)

        right_tree = ttk.Treeview(right_frame, columns=("Name", "Availability", "Notes"), show="headings")
        right_tree.heading("Name", text="Name")
        right_tree.column("Name", width=200, anchor="w", stretch=True)
        right_tree.heading("Availability", text="Available Days")
        right_tree.column("Availability", width=200, anchor="w", stretch=True)
        right_tree.heading("Notes", text="Notes")
        right_tree.column("Notes", width=200, anchor="w", stretch=True)
        right_tree.grid(row=0, column=0, sticky="nsew")

        right_scroll = ttk.Scrollbar(right_frame, orient="vertical", command=right_tree.yview)
        right_tree.configure(yscrollcommand=right_scroll.set)
        right_scroll.grid(row=0, column=1, sticky="ns")

        # Frame to hold name + checkboxes (left) and notes (right)
        entry_frame = tk.Frame(top)
        entry_frame.grid(row=1, column=0, columnspan=2, pady=5, padx=10, sticky="nsew")
        entry_frame.columnconfigure(0, weight=1)
        entry_frame.columnconfigure(1, weight=1)
        top.rowconfigure(1, weight=0)
        top.columnconfigure(0, weight=1)
        top.columnconfigure(1, weight=1)

        # Name entry (left top)
        name_label = tk.Label(entry_frame, text="Name:")
        name_label.grid(row=0, column=0, sticky="w", padx=5)
        name_entry = tk.Entry(entry_frame, width=30)
        name_entry.grid(row=1, column=0, sticky="nw", padx=8)

        # Checkboxes (row below name, same left column)
        check_vars = {day: tk.BooleanVar() for day in self.days}
        check_frame = tk.Frame(entry_frame)
        check_frame.grid(row=1, column=0, sticky="nw", pady=(25, 0))
        for day in self.days:
            tk.Checkbutton(check_frame, text=day, variable=check_vars[day]).pack(side=tk.LEFT, padx=3)

        # Notes entry (right top)
        notes_label = tk.Label(entry_frame, text="Notes:")
        notes_label.grid(row=0, column=1, sticky="w", padx=(20, 0))
        notes_entry = tk.Text(entry_frame, width=50, height=4, wrap="word")
        notes_entry.grid(row=1, column=1, sticky="w", padx=(20, 0))
        notes_entry.insert("1.0", "Enter notes here (optional)")

        # Treeview row styling for odd rows
        style = ttk.Style()
        style.configure("Treeview", rowheight=25)
        style.map("Treeview", background=[("selected", "#347083")])
        left_tree.tag_configure("oddrow", background="#AAC1DC")
        right_tree.tag_configure("oddrow", background="#AAC1DC")

        def load_employees():
            left_tree.delete(*left_tree.get_children())
            right_tree.delete(*right_tree.get_children())
            employee_row_map.clear()
            all_employees = []

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                name = row[0] if row[0] else ""
                available_days = [self.days[i] for i, val in enumerate(row[1:8]) if val == "Yes"]

                # Ensure notes are read from column 9 (index 8)
                notes = ""
                if len(row) > 8:
                    raw_note = row[8]
                    if isinstance(raw_note, str):
                        notes = raw_note.strip()
                    elif isinstance(raw_note, float) and not math.isnan(raw_note):
                        notes = str(raw_note)  # convert number notes to string
                    elif raw_note is not None:
                        notes = str(raw_note)

                all_employees.append((name, ", ".join(available_days), notes, idx))

            midpoint = len(all_employees) // 2
            left_employees[:] = all_employees[:midpoint]
            right_employees[:] = all_employees[midpoint:]

            for i, (name, days, notes, row_idx) in enumerate(left_employees):
                tag = "oddrow" if i % 2 == 1 else ""
                left_tree.insert("", "end", values=(name, days, notes), tags=(tag,))
                employee_row_map[name] = row_idx

            for i, (name, days, notes, row_idx) in enumerate(right_employees):
                tag = "oddrow" if i % 2 == 1 else ""
                right_tree.insert("", "end", values=(name, days, notes), tags=(tag,))
                employee_row_map[name] = row_idx

        def clear_selection():
            for tree in [left_tree, right_tree]:
                tree.selection_remove(tree.selection())  # Deselect all
            name_entry.delete(0, tk.END)
            for var in check_vars.values():
                var.set(False)
            notes_entry.delete(1.0, tk.END)  # Clear notes as well

        def on_select(event, tree):
            selected = tree.selection()
            if not selected:
                return
            name = tree.item(selected[0])["values"][0]
            row_idx = employee_row_map.get(name)
            if row_idx:
                row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                name_entry.delete(0, tk.END)
                name_entry.insert(0, row[0])

                # Set the availability checkboxes based on the row data
                for i, day in enumerate(self.days):
                    check_vars[day].set(row[i + 1] == "Yes")

                # Set the notes entry field with the appropriate value from the row
                notes = row[8] if len(row) > 8 and row[8] is not None else ""  # Ensure notes are retrieved from the correct column
                notes_entry.delete(1.0, tk.END)  # Clear any existing text
                notes_entry.insert(tk.END, notes)  # Insert the notes into the Text widget

            # Clear selection in the other tree
            if tree == left_tree:
                right_tree.selection_remove(right_tree.selection())
            else:
                left_tree.selection_remove(left_tree.selection())

        def add_employee():
            name = name_entry.get().strip()
            notes = notes_entry.get("1.0", "end-1c").strip()
            if not name:
                messagebox.showerror("Error", "Name cannot be empty.")
                return
            if any(row[0].value == name for row in sheet.iter_rows(min_row=2)):
                messagebox.showerror("Error", "Employee already exists.")
                return
            new_row = [name] + [("Yes" if check_vars[day].get() else "") for day in self.days] + [notes]
            sheet.append(new_row)
            wb.save(file_path)
            load_employees()
            clear_selection()

        def delete_selected():
            selected = left_tree.selection() or right_tree.selection()
            if not selected:
                return
            tree = left_tree if selected[0] in left_tree.selection() else right_tree
            name_to_delete = tree.item(selected[0])["values"][0]

            # Store deleted employee to the recently_deleted.xlsx file
            deleted_employees = []
            for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if row[0].value == name_to_delete:
                    # Store the deleted employee's name and data
                    deleted_employees.append((name_to_delete, row[1:]))  
                    sheet.delete_rows(idx, 1)  # Delete the employee from the sheet
                    break

            if deleted_employees:
                # Load the recently_deleted.xlsx file
                deleted_wb = load_workbook(deleted_file_path)
                deleted_sheet = deleted_wb.active

                # Find the last row of the deleted employees file (to append)
                last_row = deleted_sheet.max_row + 1

                # Append the deleted employees to the deleted file
                for employee in deleted_employees:
                    deleted_sheet.append([employee[0]] + [cell.value for cell in employee[1]])

                # Save the updated deleted employees workbook
                deleted_wb.save(deleted_file_path)

                # Save the updated main file (file_path) after deletion
                wb.save(file_path)

                load_employees()  # Reload employee data
                clear_selection()  # Clear selection in the UI

        def undo_delete():
            deleted_wb = load_workbook(deleted_file_path)
            deleted_sheet = deleted_wb.active

            # Start from the bottom of the sheet and search for the last non-empty row
            for row_num in range(deleted_sheet.max_row, 0, -1):
                name_cell = deleted_sheet.cell(row=row_num, column=1)
                if name_cell.value:  # Only proceed if name cell has data
                    name = name_cell.value
                    available_days = [
                        deleted_sheet.cell(row=row_num, column=col).value for col in range(2, len(self.days) + 2)
                    ]

                    # Check if the employee already exists in the main sheet
                    existing_names = [row[0].value for row in sheet.iter_rows(min_row=2)]
                    if name not in existing_names:
                        # Append to main sheet
                        sheet.append([name] + available_days)

                        # Delete the restored row from the deleted sheet
                        deleted_sheet.delete_rows(row_num)

                        # Save both workbooks
                        deleted_wb.save(deleted_file_path)
                        wb.save(file_path)

                        # Update UI
                        load_employees()
                        clear_selection()
                    else:
                        messagebox.showinfo("Info", "Employee already exists in the main sheet.")
                    return

            # If no non-empty deleted rows were found
            messagebox.showinfo("Info", "No deleted employees found to undo.")

        def update_employee():
            name = name_entry.get().strip()
            if not name:
                messagebox.showerror("Error", "Name cannot be empty.")
                return
            row_idx = employee_row_map.get(name)
            if not row_idx:
                messagebox.showerror("Error", "Employee not found.")
                return
            for i, day in enumerate(self.days):
                value = "Yes" if check_vars[day].get() else ""
                sheet.cell(row=row_idx, column=i + 2).value = value
            notes = notes_entry.get("1.0", "end-1c").strip()
            print("Employee row map:", employee_row_map)
            sheet.cell(row=row_idx, column=9).value = notes
            wb.save(file_path)
            load_employees()
            clear_selection()

        def on_click_inside(event):
            widget = event.widget
            # Don't clear if clicking on Treeviews or form widgets
            allowed_widgets = [left_tree, right_tree, name_entry, notes_entry]
            allowed_widget_classes = (tk.Entry, tk.Text, tk.Checkbutton, tk.Button)

            if any(str(widget).startswith(str(w)) for w in allowed_widgets) or isinstance(widget, allowed_widget_classes):
                return  # Clicked inside an allowed widget — do nothing

            # Otherwise, clear the selection
            clear_selection()

        # Buttons
        btn_frame = tk.Frame(top)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=5)
        tk.Button(btn_frame, text="Add Employee", command=add_employee).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Update Employee", command=update_employee).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Delete Employee", command=delete_selected).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Undo Delete", command=undo_delete).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Close", command=top.destroy).pack(side=tk.LEFT, padx=5)

        left_tree.bind("<<TreeviewSelect>>", lambda e: on_select(e, left_tree))
        right_tree.bind("<<TreeviewSelect>>", lambda e: on_select(e, right_tree))

        # Bind clicking outside to clear selection
        top.bind("<Button-1>", on_click_inside)

        # Load existing employees
        load_employees()
if __name__ == "__main__":
    app = ScheduleWindow(root)
    root.mainloop()
    print("Application closed.")